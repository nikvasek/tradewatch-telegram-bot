import pandas as pd
import os
import glob
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime
import config

# –ü—Ä–æ–≤–µ—Ä—è–µ–º –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç—å Selenium –∏ –≤—ã–±–∏—Ä–∞–µ–º —Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤—É—é—â–∏–π –º–æ–¥—É–ª—å
try:
    from selenium import webdriver
    SELENIUM_AVAILABLE = True
    from tradewatch_login import process_supplier_file_with_tradewatch
    print("‚úÖ Excel processor: Selenium –¥–æ—Å—Ç—É–ø–µ–Ω")
except ImportError:
    SELENIUM_AVAILABLE = False
    from tradewatch_fallback import download_from_tradewatch
    print("‚ùå Excel processor: Selenium –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω - fallback —Ä–µ–∂–∏–º")

def format_ean_to_13_digits(ean_value):
    """
    –ü—Ä–∏–≤–æ–¥–∏—Ç EAN –∫ —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω–æ–º—É 13-—Ü–∏—Ñ—Ä–æ–≤–æ–º—É —Ñ–æ—Ä–º–∞—Ç—É
    """
    if pd.isna(ean_value):
        return None
    
    try:
        # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ —Å—Ç—Ä–æ–∫—É –∏ —É–¥–∞–ª—è–µ–º –ø—Ä–æ–±–µ–ª—ã
        ean_str = str(ean_value).strip()
        
        # –ï—Å–ª–∏ –ø–æ–ª—É—á–∏–ª–æ—Å—å –ø—É—Å—Ç–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ, –≤–æ–∑–≤—Ä–∞—â–∞–µ–º None
        if not ean_str:
            return None
        
        # –û—Å–æ–±–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –¥–ª—è –Ω–∞—É—á–Ω–æ–π –Ω–æ—Ç–∞—Ü–∏–∏
        if 'E' in ean_str.upper() or 'e' in ean_str:
            try:
                # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º —á–µ—Ä–µ–∑ float –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ –Ω–∞—É—á–Ω–æ–π –Ω–æ—Ç–∞—Ü–∏–∏
                ean_float = float(ean_str)
                ean_str = str(int(ean_float))
            except:
                pass
        
        # –£–¥–∞–ª—è–µ–º –≤—Å–µ –Ω–µ—á–∏—Å–ª–æ–≤—ã–µ —Å–∏–º–≤–æ–ª—ã
        ean_digits = ''.join(char for char in ean_str if char.isdigit())
        
        # –ï—Å–ª–∏ –ø–æ–ª—É—á–∏–ª–æ—Å—å –ø—É—Å—Ç–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ, –≤–æ–∑–≤—Ä–∞—â–∞–µ–º None
        if not ean_digits:
            return None
            
        # –û–±—Ä–µ–∑–∞–µ–º –¥–æ 13 —Ü–∏—Ñ—Ä –µ—Å–ª–∏ –±–æ–ª—å—à–µ
        if len(ean_digits) > 13:
            ean_digits = ean_digits[:13]
            
        # –î–æ–ø–æ–ª–Ω—è–µ–º –≤–µ–¥—É—â–∏–º–∏ –Ω—É–ª—è–º–∏ –¥–æ 13 —Ü–∏—Ñ—Ä
        ean_formatted = ean_digits.zfill(13)
        
        return ean_formatted
        
    except Exception as e:
        print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–∏ EAN {ean_value}: {str(e)}")
        return None

def create_hyperlinks(df):
    """
    –°–æ–∑–¥–∞–µ—Ç –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏ –≤ –∫–æ–ª–æ–Ω–∫–µ Link –Ω–∞ –æ—Å–Ω–æ–≤–µ –Ω–æ–º–µ—Ä–æ–≤ –∏–∑ –∫–æ–ª–æ–Ω–æ–∫ –º–∏–Ω–∏–º–∞–ª—å–Ω–æ–π —Ü–µ–Ω—ã
    """
    # –ò—â–µ–º –∫–æ–ª–æ–Ω–∫—É —Å –Ω–æ–º–µ—Ä–∞–º–∏ –¥–ª—è —Å—Å—ã–ª–æ–∫
    link_number_column = None
    for col in config.LINK_NUMBER_COLUMNS:
        if col in df.columns:
            link_number_column = col
            break
    
    if link_number_column is None:
        print("–ù–µ –Ω–∞–π–¥–µ–Ω–∞ –∫–æ–ª–æ–Ω–∫–∞ —Å –Ω–æ–º–µ—Ä–∞–º–∏ –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è —Å—Å—ã–ª–æ–∫")
        return df
    
    # –°–æ–∑–¥–∞–µ–º –∏–ª–∏ –æ–±–Ω–æ–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É Link
    if 'Link' not in df.columns:
        df['Link'] = ''
    
    # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –∫–æ–ª–æ–Ω–∫—É Link –≤ —Å—Ç—Ä–æ–∫–æ–≤—ã–π —Ç–∏–ø
    df['Link'] = df['Link'].astype(str)
    
    # –°–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
    for index, row in df.iterrows():
        link_number = row[link_number_column]
        if pd.notna(link_number) and str(link_number).strip():
            # –ò–∑–≤–ª–µ–∫–∞–µ–º –Ω–æ–º–µ—Ä –∏–∑ —Å—Å—ã–ª–∫–∏ –µ—Å–ª–∏ —ç—Ç–æ —É–∂–µ —Å—Å—ã–ª–∫–∞
            if str(link_number).startswith('http'):
                # –ï—Å–ª–∏ —É–∂–µ –µ—Å—Ç—å —Å—Å—ã–ª–∫–∞, –∏–∑–≤–ª–µ–∫–∞–µ–º –Ω–æ–º–µ—Ä
                try:
                    number = str(link_number).split('/')[-1]
                    # –£–±–∏—Ä–∞–µ–º .0 –∏–∑ –Ω–æ–º–µ—Ä–∞ –µ—Å–ª–∏ –æ–Ω –µ—Å—Ç—å
                    if number.endswith('.0'):
                        number = number[:-2]
                    df.at[index, 'Link'] = f"{config.ALLEGRO_BASE_URL}{number}"
                except:
                    df.at[index, 'Link'] = str(link_number)
            else:
                # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é —Å—Å—ã–ª–∫—É
                # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ —á–∏—Å–ª–æ –∏ –æ–±—Ä–∞—Ç–Ω–æ –≤ —Å—Ç—Ä–æ–∫—É –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è .0
                try:
                    # –ü—ã—Ç–∞–µ–º—Å—è –∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å –≤ —á–∏—Å–ª–æ
                    clean_number = str(int(float(link_number)))
                except (ValueError, TypeError):
                    # –ï—Å–ª–∏ –Ω–µ —É–¥–∞–µ—Ç—Å—è –∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä–æ–≤–∞—Ç—å, –ø—Ä–æ—Å—Ç–æ —É–¥–∞–ª—è–µ–º .0
                    clean_number = str(link_number).rstrip('.0')
                    if clean_number == '':
                        clean_number = '0'
                
                df.at[index, 'Link'] = f"{config.ALLEGRO_BASE_URL}{clean_number}"
    
    print(f"–°–æ–∑–¥–∞–Ω—ã –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏ –Ω–∞ –æ—Å–Ω–æ–≤–µ –∫–æ–ª–æ–Ω–∫–∏: {link_number_column}")
    return df

def create_product_links(df):
    """
    –°–æ–∑–¥–∞–µ—Ç –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏ –≤ –∫–æ–ª–æ–Ω–∫–µ Product Link –Ω–∞ –æ—Å–Ω–æ–≤–µ EAN –∫–æ–¥–æ–≤
    –¢–û–õ–¨–ö–û –µ—Å–ª–∏ –≤ DataFrame —É–∂–µ –µ—Å—Ç—å –∫–æ–ª–æ–Ω–∫–∞ Product Link
    """
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –∫–æ–ª–æ–Ω–∫–∏ EAN
    if 'EAN' not in df.columns:
        print("–ö–æ–ª–æ–Ω–∫–∞ EAN –Ω–µ –Ω–∞–π–¥–µ–Ω–∞, –ø—Ä–æ–ø—É—Å–∫–∞–µ–º —Å–æ–∑–¥–∞–Ω–∏–µ Product Link")
        return df
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –∫–æ–ª–æ–Ω–∫–∏ Product Link –≤ –∏—Å—Ö–æ–¥–Ω—ã—Ö –¥–∞–Ω–Ω—ã—Ö
    if 'Product Link' not in df.columns:
        print("–ö–æ–ª–æ–Ω–∫–∞ Product Link –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ –≤ –ø—Ä–∞–π—Å–µ –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞, –ø—Ä–æ–ø—É—Å–∫–∞–µ–º —Å–æ–∑–¥–∞–Ω–∏–µ –≥–∏–ø–µ—Ä—Å—Å—ã–ª–æ–∫")
        return df
    
    print("–ù–∞–π–¥–µ–Ω–∞ –∫–æ–ª–æ–Ω–∫–∞ Product Link –≤ –ø—Ä–∞–π—Å–µ –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞, —Å–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏...")
    
    # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –∫–æ–ª–æ–Ω–∫—É Product Link –≤ —Å—Ç—Ä–æ–∫–æ–≤—ã–π —Ç–∏–ø
    df['Product Link'] = df['Product Link'].astype(str)
    
    # –°–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏ –Ω–∞ –æ—Å–Ω–æ–≤–µ EAN
    for index, row in df.iterrows():
        ean_value = row['EAN']
        if pd.notna(ean_value) and str(ean_value).strip():
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º EAN –≤ 13-—Ü–∏—Ñ—Ä–æ–≤–æ–π —Ñ–æ—Ä–º–∞—Ç
            formatted_ean = format_ean_to_13_digits(ean_value)
            if formatted_ean:
                # –°–æ–∑–¥–∞–µ–º URL —Å EAN
                product_url = f"https://api.qogita.com/variants/link/{formatted_ean}/"
                df.at[index, 'Product Link'] = product_url
    
    print("–°–æ–∑–¥–∞–Ω—ã Product Link –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏ –Ω–∞ –æ—Å–Ω–æ–≤–µ EAN –∫–æ–¥–æ–≤")
    return df

def calculate_profit_and_roi(df):
    """
    –î–æ–±–∞–≤–ª—è–µ—Ç –∫–æ–ª–æ–Ω–∫–∏ Profit –∏ ROI –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ –≤ Excel —Ñ–æ—Ä–º—É–ª–∞–º–∏
    """
    # –î–æ–±–∞–≤–ª—è–µ–º –ø—É—Å—Ç—ã–µ –∫–æ–ª–æ–Ω–∫–∏ –¥–ª—è Profit –∏ ROI
    # –§–æ—Ä–º—É–ª—ã –±—É–¥—É—Ç –¥–æ–±–∞–≤–ª–µ–Ω—ã –ø—Ä–∏ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–∏ Excel
    if 'Profit' not in df.columns:
        df['Profit'] = None
    if 'ROI' not in df.columns:
        df['ROI'] = None
    
    print("–î–æ–±–∞–≤–ª–µ–Ω—ã –∫–æ–ª–æ–Ω–∫–∏ Profit –∏ ROI –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ —Ñ–æ—Ä–º—É–ª–∞–º–∏ Excel")
    return df

def merge_excel_files_by_ean_with_calculations(directory_path='.'):
    """
    –û–±—ä–µ–¥–∏–Ω—è–µ—Ç —Ñ–∞–π–ª—ã Excel –ø–æ EAN –∫–æ–¥—É —Å —Ä–∞—Å—á–µ—Ç–æ–º –ø—Ä–∏–±—ã–ª–∏ –∏ ROI.
    """
    
    # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ —Ñ–∞–π–ª—ã TradeWatch
    tradewatch_files = glob.glob(os.path.join(directory_path, config.TRADEWATCH_FILE_PATTERN))
    
    if not tradewatch_files:
        print("–ù–µ –Ω–∞–π–¥–µ–Ω—ã —Ñ–∞–π–ª—ã, –Ω–∞—á–∏–Ω–∞—é—â–∏–µ—Å—è —Å 'TradeWatch'")
        return
    
    print(f"–ù–∞–π–¥–µ–Ω–æ —Ñ–∞–π–ª–æ–≤ TradeWatch: {len(tradewatch_files)}")
    for file in tradewatch_files:
        print(f"  - {os.path.basename(file)}")
    
    # –°–æ–±–∏—Ä–∞–µ–º –≤—Å–µ EAN –∫–æ–¥—ã –∏–∑ —Ñ–∞–π–ª–æ–≤ TradeWatch
    all_ean_data = []
    
    for file_path in tradewatch_files:
        try:
            print(f"\n–û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª: {os.path.basename(file_path)}")
            
            # –ß–∏—Ç–∞–µ–º –ª–∏—Å—Ç "Produkty wg EAN"
            df = pd.read_excel(file_path, sheet_name=config.TRADEWATCH_SHEET_NAME)
            
            # –ò—â–µ–º –∫–æ–ª–æ–Ω–∫—É EAN —Å—Ä–µ–¥–∏ –≤–æ–∑–º–æ–∂–Ω—ã—Ö –≤–∞—Ä–∏–∞–Ω—Ç–æ–≤ –Ω–∞–∑–≤–∞–Ω–∏–π
            ean_column = None
            possible_ean_columns = ['EAN', 'ean', 'EAN13', 'EAN-13', 'GTIN', 'gtin', '–ö–æ–¥ EAN', 'EAN –∫–æ–¥']
            
            for col in possible_ean_columns:
                if col in df.columns:
                    ean_column = col
                    break
            
            if ean_column is None:
                print(f"  ‚ö†Ô∏è –ö–æ–ª–æ–Ω–∫–∞ EAN –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ –≤ —Ñ–∞–π–ª–µ {file_path}")
                print(f"  üìã –î–æ—Å—Ç—É–ø–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏: {list(df.columns)}")
                continue
            
            print(f"  ‚úÖ –ù–∞–π–¥–µ–Ω–∞ –∫–æ–ª–æ–Ω–∫–∞ EAN: '{ean_column}'")
            
            # –ü–µ—Ä–µ–∏–º–µ–Ω–æ–≤—ã–≤–∞–µ–º –∫–æ–ª–æ–Ω–∫—É –≤ —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω–æ–µ –∏–º—è 'EAN'
            if ean_column != 'EAN':
                df = df.rename(columns={ean_column: 'EAN'})
            
            # –î–æ–±–∞–≤–ª—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ–± –∏—Å—Ç–æ—á–Ω–∏–∫–µ
            df['source_file'] = os.path.basename(file_path)
            
            # –§–∏–ª—å—Ç—Ä—É–µ–º —Ç–æ–ª—å–∫–æ —Å—Ç—Ä–æ–∫–∏ —Å –≤–∞–ª–∏–¥–Ω—ã–º–∏ EAN
            df_clean = df[df['EAN'].notna()].copy()
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º EAN –≤ 13-—Ü–∏—Ñ—Ä–æ–≤–æ–π —Ñ–æ—Ä–º–∞—Ç
            df_clean['EAN'] = df_clean['EAN'].apply(format_ean_to_13_digits)
            
            # –£–±–∏—Ä–∞–µ–º —Å—Ç—Ä–æ–∫–∏ —Å –Ω–µ–≤–∞–ª–∏–¥–Ω—ã–º–∏ EAN
            df_clean = df_clean[df_clean['EAN'].notna()].copy()
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º EAN –≤ 13-—Ü–∏—Ñ—Ä–æ–≤–æ–º —Ñ–æ—Ä–º–∞—Ç–µ
            df_clean['EAN'] = df_clean['EAN'].apply(format_ean_to_13_digits)
            
            all_ean_data.append(df_clean)
            print(f"  –ù–∞–π–¥–µ–Ω–æ EAN –∫–æ–¥–æ–≤: {len(df_clean)}")
            
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ —Ñ–∞–π–ª–∞ {file_path}: {str(e)}")
    
    if not all_ean_data:
        print("–ù–µ —É–¥–∞–ª–æ—Å—å –∏–∑–≤–ª–µ—á—å –¥–∞–Ω–Ω—ã–µ –∏–∑ —Ñ–∞–π–ª–æ–≤ TradeWatch")
        return
    
    # –û–±—ä–µ–¥–∏–Ω—è–µ–º –≤—Å–µ –¥–∞–Ω–Ω—ã–µ TradeWatch
    combined_tradewatch = pd.concat(all_ean_data, ignore_index=True)
    print(f"\n–í—Å–µ–≥–æ —É–Ω–∏–∫–∞–ª—å–Ω—ã—Ö EAN –∫–æ–¥–æ–≤ –∏–∑ TradeWatch: {combined_tradewatch['EAN'].nunique()}")
    
    # –ù–∞—Ö–æ–¥–∏–º –¥—Ä—É–≥–∏–µ Excel —Ñ–∞–π–ª—ã –¥–ª—è –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏—è
    all_excel_files = glob.glob(os.path.join(directory_path, "*.xlsx"))
    
    # –ò—Å–∫–ª—é—á–∞–µ–º —Ñ–∞–π–ª—ã TradeWatch, –≤—Ä–µ–º–µ–Ω–Ω—ã–µ —Ñ–∞–π–ª—ã Excel –∏ —Ñ–∞–π–ª—ã —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤
    other_files = []
    
    for f in all_excel_files:
        filename = os.path.basename(f)
        if not any(filename.startswith(prefix) for prefix in config.EXCLUDED_FILE_PREFIXES):
            other_files.append(f)
    
    if not other_files:
        print("–ù–µ –Ω–∞–π–¥–µ–Ω—ã –¥—Ä—É–≥–∏–µ Excel —Ñ–∞–π–ª—ã –¥–ª—è –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏—è")
        # –†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º –ø—Ä–∏–±—ã–ª—å –∏ ROI –¥–ª—è –¥–∞–Ω–Ω—ã—Ö TradeWatch
        combined_tradewatch = calculate_profit_and_roi(combined_tradewatch)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É Price PL
        combined_tradewatch = add_price_pl_column(combined_tradewatch)
        
        # –°–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        combined_tradewatch = create_hyperlinks(combined_tradewatch)
        
        # –°–æ–∑–¥–∞–µ–º Product Link –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        combined_tradewatch = create_product_links(combined_tradewatch)
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ç–æ–ª—å–∫–æ –¥–∞–Ω–Ω—ã–µ TradeWatch
        output_file = os.path.join(directory_path, config.OUTPUT_FILE_MERGED_TRADEWATCH)
        save_formatted_excel(combined_tradewatch, output_file)
        print(f"–°–æ—Ö—Ä–∞–Ω–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ TradeWatch –≤ —Ñ–∞–π–ª: {output_file}")
        return
    
    print(f"\n–ù–∞–π–¥–µ–Ω–æ –¥—Ä—É–≥–∏—Ö Excel —Ñ–∞–π–ª–æ–≤: {len(other_files)}")
    for file in other_files:
        print(f"  - {os.path.basename(file)}")
    
    # –û–±—ä–µ–¥–∏–Ω—è–µ–º —Å –¥—Ä—É–≥–∏–º–∏ —Ñ–∞–π–ª–∞–º–∏
    merged_results = []
    
    for other_file in other_files:
        try:
            print(f"\n–û–±—ä–µ–¥–∏–Ω—è–µ–º —Å —Ñ–∞–π–ª–æ–º: {os.path.basename(other_file)}")
            
            # –ß–∏—Ç–∞–µ–º —Ñ–∞–π–ª (–ø—Ä–æ–±—É–µ–º –ø–µ—Ä–≤—ã–π –ª–∏—Å—Ç)
            other_df = pd.read_excel(other_file)
            
            # –ò—â–µ–º –∫–æ–ª–æ–Ω–∫—É —Å GTIN –∏–ª–∏ EAN
            gtin_column = None
            
            for col in config.POSSIBLE_GTIN_COLUMNS:
                if col in other_df.columns:
                    gtin_column = col
                    break
            
            if gtin_column is None:
                print(f"  –ù–µ –Ω–∞–π–¥–µ–Ω–∞ –∫–æ–ª–æ–Ω–∫–∞ GTIN/EAN –≤ —Ñ–∞–π–ª–µ {other_file}")
                print(f"  –î–æ—Å—Ç—É–ø–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏: {list(other_df.columns)}")
                continue
            
            # –û—á–∏—â–∞–µ–º –¥–∞–Ω–Ω—ã–µ GTIN
            other_df_clean = other_df[other_df[gtin_column].notna()].copy()
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º GTIN –≤ 13-—Ü–∏—Ñ—Ä–æ–≤–æ–π —Ñ–æ—Ä–º–∞—Ç
            other_df_clean[gtin_column] = other_df_clean[gtin_column].apply(format_ean_to_13_digits)
            
            # –£–±–∏—Ä–∞–µ–º —Å—Ç—Ä–æ–∫–∏ —Å –Ω–µ–≤–∞–ª–∏–¥–Ω—ã–º–∏ GTIN
            other_df_clean = other_df_clean[other_df_clean[gtin_column].notna()].copy()
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º GTIN –≤ 13-—Ü–∏—Ñ—Ä–æ–≤–æ–º —Ñ–æ—Ä–º–∞—Ç–µ
            other_df_clean[gtin_column] = other_df_clean[gtin_column].apply(format_ean_to_13_digits)
            
            # –û–±—ä–µ–¥–∏–Ω—è–µ–º –ø–æ EAN/GTIN
            merged = pd.merge(
                combined_tradewatch,
                other_df_clean,
                left_on='EAN',
                right_on=gtin_column,
                how='inner'  # –û—Å—Ç–∞–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ —Å–æ–≤–ø–∞–¥–∞—é—â–∏–µ EAN
            )
            
            if len(merged) > 0:
                merged['merged_with'] = os.path.basename(other_file)
                merged_results.append(merged)
                print(f"  –ù–∞–π–¥–µ–Ω–æ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–π: {len(merged)}")
            else:
                print(f"  –°–æ–≤–ø–∞–¥–µ–Ω–∏–π –Ω–µ –Ω–∞–π–¥–µ–Ω–æ")
                
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏–∏ —Å —Ñ–∞–π–ª–æ–º {other_file}: {str(e)}")
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
    if merged_results:
        # –û–±—ä–µ–¥–∏–Ω—è–µ–º –≤—Å–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
        final_result = pd.concat(merged_results, ignore_index=True)
        
        # –†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º –ø—Ä–∏–±—ã–ª—å –∏ ROI
        final_result = calculate_profit_and_roi(final_result)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É Price PL
        final_result = add_price_pl_column(final_result)
        
        # –°–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        final_result = create_hyperlinks(final_result)
        
        # –°–æ–∑–¥–∞–µ–º Product Link –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        final_result = create_product_links(final_result)
        
        # –£–ø–æ—Ä—è–¥–æ—á–∏–≤–∞–µ–º –∫–æ–ª–æ–Ω–∫–∏ –≤ –Ω—É–∂–Ω–æ–º –ø–æ—Ä—è–¥–∫–µ –∏ –æ—Å—Ç–∞–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ –Ω—É–∂–Ω—ã–µ
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –¥–æ—Å—Ç—É–ø–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏ –∏–∑ –∂–µ–ª–∞–µ–º–æ–≥–æ –ø–æ—Ä—è–¥–∫–∞ (–æ—Å—Ç–∞–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ –∏—Ö)
        available_ordered_columns = [col for col in config.DESIRED_COLUMN_ORDER if col in final_result.columns]
        
        # –§–æ—Ä–º–∏—Ä—É–µ–º —Ñ–∏–Ω–∞–ª—å–Ω—ã–π –ø–æ—Ä—è–¥–æ–∫ –∫–æ–ª–æ–Ω–æ–∫ (—Ç–æ–ª—å–∫–æ —É–∫–∞–∑–∞–Ω–Ω—ã–µ –≤ DESIRED_COLUMN_ORDER)
        final_column_order = available_ordered_columns
        
        # –ü–µ—Ä–µ—É–ø–æ—Ä—è–¥–æ—á–∏–≤–∞–µ–º DataFrame –∏ –æ—Å—Ç–∞–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ –Ω—É–∂–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏
        final_result = final_result[final_column_order]
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ —Ñ–∞–π–ª —Å —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ–º
        output_file = os.path.join(directory_path, config.OUTPUT_FILE_WITH_CALCULATIONS)
        save_formatted_excel(final_result, output_file)
        
        # –ü–æ–¥–≥–æ—Ç–∞–≤–ª–∏–≤–∞–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É
        stats = {
            'total_rows': len(final_result),
            'unique_ean': final_result['EAN'].nunique() if 'EAN' in final_result.columns else 0,
            'files_processed': len(other_files),
            'output_file': output_file
        }
        
        print(f"\n–†–µ–∑—É–ª—å—Ç–∞—Ç —Å–æ—Ö—Ä–∞–Ω–µ–Ω –≤ —Ñ–∞–π–ª: {output_file}")
        print(f"–í—Å–µ–≥–æ —Å—Ç—Ä–æ–∫ –≤ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–µ: {stats['total_rows']}")
        print(f"–£–Ω–∏–∫–∞–ª—å–Ω—ã—Ö EAN –∫–æ–¥–æ–≤: {stats['unique_ean']}")
        print(f"–ü–æ—Ä—è–¥–æ–∫ –∫–æ–ª–æ–Ω–æ–∫: {final_column_order}")
        
        return stats
    else:
        print("\n–°–æ–≤–ø–∞–¥–µ–Ω–∏–π –Ω–µ –Ω–∞–π–¥–µ–Ω–æ, —Å–æ—Ö—Ä–∞–Ω—è–µ–º —Ç–æ–ª—å–∫–æ –¥–∞–Ω–Ω—ã–µ TradeWatch")
        # –†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º –ø—Ä–∏–±—ã–ª—å –∏ ROI –¥–ª—è –¥–∞–Ω–Ω—ã—Ö TradeWatch
        combined_tradewatch = calculate_profit_and_roi(combined_tradewatch)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É Price PL
        combined_tradewatch = add_price_pl_column(combined_tradewatch)
        
        # –°–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        combined_tradewatch = create_hyperlinks(combined_tradewatch)
        
        # –°–æ–∑–¥–∞–µ–º Product Link –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        combined_tradewatch = create_product_links(combined_tradewatch)
        
        output_file = os.path.join(directory_path, config.OUTPUT_FILE_TRADEWATCH_ONLY)
        save_formatted_excel(combined_tradewatch, output_file)
        print(f"–î–∞–Ω–Ω—ã–µ TradeWatch —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã –≤ —Ñ–∞–π–ª: {output_file}")

def save_formatted_excel(df, output_file):
    """
    –°–æ—Ö—Ä–∞–Ω—è–µ—Ç DataFrame –≤ Excel —Å —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ–º —Å–æ–≥–ª–∞—Å–Ω–æ –∫–æ–Ω—Ñ–∏–≥—É—Ä–∞—Ü–∏–∏
    """
    try:
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º DataFrame –≤ —Ñ–∞–π–ª
        df.to_excel(output_file, index=False, startrow=config.EXCEL_TO_EXCEL_STARTROW)
        
        # –ó–∞–≥—Ä—É–∂–∞–µ–º workbook –¥–ª—è —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
        wb = load_workbook(output_file)
        ws = wb.active
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏ –≤ –ø–µ—Ä–≤—ã–µ —Å—Ç—Ä–æ–∫–∏
        for row_num, row_data in config.TITLE_ROWS.items():
            for col_letter, value in row_data.items():
                if value is not None:
                    cell = ws[f"{col_letter}{row_num}"]
                    cell.value = value
                    
                    # –ü—Ä–∏–º–µ–Ω—è–µ–º —Ü–≤–µ—Ç–∞ –∏ —Å—Ç–∏–ª–∏ –¥–ª—è –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
                    if col_letter in ['D', 'I']:  # –ö–æ–ª–æ–Ω–∫–∏ —Å –Ω–∞–∑–≤–∞–Ω–∏—è–º–∏
                        cell.font = Font(
                            name=config.TITLE_FONT_SETTINGS['name'],
                            size=config.TITLE_FONT_SETTINGS['size'],
                            bold=config.TITLE_FONT_SETTINGS['bold'],
                            color=config.TITLE_LABEL_FONT_COLOR
                        )
                    elif col_letter in ['G', 'M']:  # –ö–æ–ª–æ–Ω–∫–∏ —Å–æ –∑–Ω–∞—á–µ–Ω–∏—è–º–∏
                        cell.font = Font(
                            name=config.TITLE_FONT_SETTINGS['name'],
                            size=config.TITLE_FONT_SETTINGS['size'],
                            bold=config.TITLE_FONT_SETTINGS['bold'],
                            color=config.TITLE_VALUE_FONT_COLOR
                        )
        
        # –ù–∞—Å—Ç—Ä–∞–∏–≤–∞–µ–º –∑–∞–≥–æ–ª–æ–≤–∫–∏ —Ç–∞–±–ª–∏—Ü—ã
        header_row = config.EXCEL_HEADER_ROW_NUM
        ws.row_dimensions[header_row].height = config.EXCEL_HEADER_ROW_HEIGHT
        
        # –°—Ç–∏–ª–∏ –¥–ª—è –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
        header_font = Font(
            name=config.HEADER_FONT['name'],
            size=config.HEADER_FONT['size'],
            bold=config.HEADER_FONT['bold'],
            color=config.HEADER_FONT['color']
        )
        
        header_alignment = Alignment(
            horizontal=config.HEADER_ALIGNMENT['horizontal'],
            vertical=config.HEADER_ALIGNMENT['vertical'],
            wrap_text=config.HEADER_ALIGNMENT['wrap_text']
        )
        
        header_fill = PatternFill(
            start_color=config.HEADER_FILL['start_color'],
            end_color=config.HEADER_FILL['end_color'],
            fill_type=config.HEADER_FILL['fill_type']
        )
        
        # –°—Ç–∏–ª–∏ –¥–ª—è –¥–∞–Ω–Ω—ã—Ö
        data_font = Font(
            name=config.DATA_FONT['name'],
            size=config.DATA_FONT['size']
        )
        
        # –í—ã—Ä–∞–≤–Ω–∏–≤–∞–Ω–∏–µ –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é
        data_alignment = Alignment(
            horizontal=config.DATA_ALIGNMENT['horizontal'],
            vertical=config.DATA_ALIGNMENT['vertical']
        )
        
        # –í—ã—Ä–∞–≤–Ω–∏–≤–∞–Ω–∏–µ –ø–æ –ø—Ä–∞–≤–æ–º—É –∫—Ä–∞—é
        data_alignment_right = Alignment(
            horizontal=config.DATA_ALIGNMENT_RIGHT['horizontal'],
            vertical=config.DATA_ALIGNMENT_RIGHT['vertical']
        )
        
        # –°—Ç–∏–ª—å –≥—Ä–∞–Ω–∏—Ü –¥–ª—è –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
        header_border_side = Side(
            border_style=config.HEADER_BORDER_STYLE['border_style'],
            color=config.HEADER_BORDER_STYLE['color']
        )
        header_border = Border(left=header_border_side, right=header_border_side, top=header_border_side, bottom=header_border_side)
        
        # –ü—Ä–∏–º–µ–Ω—è–µ–º —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∫ –∑–∞–≥–æ–ª–æ–≤–∫–∞–º —Ç–∞–±–ª–∏—Ü—ã
        for col_num, column in enumerate(df.columns, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = header_border
            
            # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º —à–∏—Ä–∏–Ω—É –∫–æ–ª–æ–Ω–∫–∏
            column_letter = cell.column_letter
            if column in config.WIDE_COLUMNS:
                ws.column_dimensions[column_letter].width = config.WIDE_COLUMNS[column]
            else:
                ws.column_dimensions[column_letter].width = config.DEFAULT_COLUMN_WIDTH
        
        # –ü—Ä–∏–º–µ–Ω—è–µ–º —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∫ –¥–∞–Ω–Ω—ã–º (–±–µ–∑ –≥—Ä–∞–Ω–∏—Ü)
        for row_num in range(config.EXCEL_DATA_START_ROW_NUM, len(df) + config.EXCEL_DATA_START_ROW_NUM):
            for col_num in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                
                # –ü–æ–ª—É—á–∞–µ–º –∏–º—è –∫–æ–ª–æ–Ω–∫–∏
                column_name = df.columns[col_num - 1]
                
                # –ü—Ä–∏–º–µ–Ω—è–µ–º –≤—ã—Ä–∞–≤–Ω–∏–≤–∞–Ω–∏–µ –≤ –∑–∞–≤–∏—Å–∏–º–æ—Å—Ç–∏ –æ—Ç –∫–æ–ª–æ–Ω–∫–∏
                if column_name in config.RIGHT_ALIGNED_COLUMNS:
                    cell.alignment = data_alignment_right
                else:
                    cell.alignment = data_alignment
                
                # –ü—Ä–∏–º–µ–Ω—è–µ–º —á–∏—Å–ª–æ–≤—ã–µ —Ñ–æ—Ä–º–∞—Ç—ã
                if column_name in config.PRICE_FORMAT_COLUMNS:
                    cell.number_format = config.PRICE_NUMBER_FORMAT
                elif column_name in config.ROI_FORMAT_COLUMNS:
                    cell.number_format = config.ROI_NUMBER_FORMAT
                elif column_name in config.EAN_FORMAT_COLUMNS:
                    cell.number_format = config.EAN_NUMBER_FORMAT
                
                # –î–æ–±–∞–≤–ª—è–µ–º —Ñ–æ—Ä–º—É–ª—É –¥–ª—è –∫–æ–ª–æ–Ω–∫–∏ Price PL
                if column_name == 'Price PL':
                    # –ù–∞—Ö–æ–¥–∏–º –∫–æ–ª–æ–Ω–∫—É Price
                    price_col_num = None
                    for i, col in enumerate(df.columns):
                        if col == 'Price':
                            price_col_num = i + 1
                            break
                    
                    if price_col_num:
                        # –ü–æ–ª—É—á–∞–µ–º –±—É–∫–≤—ã –∫–æ–ª–æ–Ω–æ–∫
                        from openpyxl.utils import get_column_letter
                        price_col_letter = get_column_letter(price_col_num)
                        
                        # –°–æ–∑–¥–∞–µ–º —Ñ–æ—Ä–º—É–ª—É: Price * G1 (–∫—É—Ä—Å –æ–±–º–µ–Ω–∞)
                        formula = f"={price_col_letter}{row_num}*$G$1"
                        cell.value = formula
                        
                        # –ü—Ä–∏–º–µ–Ω—è–µ–º —Ñ–æ—Ä–º–∞—Ç —á–∏—Å–ª–∞ (–±—É–¥–µ—Ç –ø—Ä–∏–º–µ–Ω–µ–Ω –ø–æ–∑–∂–µ —á–µ—Ä–µ–∑ config.PRICE_NUMBER_FORMAT)
                
                # –î–æ–±–∞–≤–ª—è–µ–º —Ñ–æ—Ä–º—É–ª—É –¥–ª—è –∫–æ–ª–æ–Ω–∫–∏ Profit
                elif column_name == 'Profit':
                    # –ù–∞—Ö–æ–¥–∏–º –Ω—É–∂–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏
                    price_pl_col_num = None
                    cena_min_col_num = None
                    
                    for i, col in enumerate(df.columns):
                        if col == 'Price PL':
                            price_pl_col_num = i + 1
                        elif col == 'Cena min.':
                            cena_min_col_num = i + 1
                    
                    if price_pl_col_num and cena_min_col_num:
                        # –ü–æ–ª—É—á–∞–µ–º –±—É–∫–≤—ã –∫–æ–ª–æ–Ω–æ–∫
                        from openpyxl.utils import get_column_letter
                        price_pl_col_letter = get_column_letter(price_pl_col_num)
                        cena_min_col_letter = get_column_letter(cena_min_col_num)
                        
                        # –°–æ–∑–¥–∞–µ–º —Ñ–æ—Ä–º—É–ª—É –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ –ø—Ä–∏–±—ã–ª–∏:
                        # Profit = (Cena min. / 1.23) - ((Cena min. * M1%) / 1.23) - Price PL - –î–æ—Å—Ç–∞–≤–∫–∞(G2) - –°—Ç–æ–∏–º–æ—Å—Ç—å Prep Center(G3)
                        formula = f"=({cena_min_col_letter}{row_num}/1.23)-(({cena_min_col_letter}{row_num}*$M$1/100)/1.23)-{price_pl_col_letter}{row_num}-$G$2-$G$3"
                        cell.value = formula
                        
                        # –ü—Ä–∏–º–µ–Ω—è–µ–º —Ñ–æ—Ä–º–∞—Ç —á–∏—Å–ª–∞ (–±—É–¥–µ—Ç –ø—Ä–∏–º–µ–Ω–µ–Ω –ø–æ–∑–∂–µ —á–µ—Ä–µ–∑ config.PRICE_NUMBER_FORMAT)
                
                # –î–æ–±–∞–≤–ª—è–µ–º —Ñ–æ—Ä–º—É–ª—É –¥–ª—è –∫–æ–ª–æ–Ω–∫–∏ ROI
                elif column_name == 'ROI':
                    # –ù–∞—Ö–æ–¥–∏–º –Ω—É–∂–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏
                    profit_col_num = None
                    cena_min_col_num = None
                    
                    for i, col in enumerate(df.columns):
                        if col == 'Profit':
                            profit_col_num = i + 1
                        elif col == 'Cena min.':
                            cena_min_col_num = i + 1
                    
                    if profit_col_num and cena_min_col_num:
                        # –ü–æ–ª—É—á–∞–µ–º –±—É–∫–≤—ã –∫–æ–ª–æ–Ω–æ–∫
                        from openpyxl.utils import get_column_letter
                        profit_col_letter = get_column_letter(profit_col_num)
                        cena_min_col_letter = get_column_letter(cena_min_col_num)
                        
                        # –°–æ–∑–¥–∞–µ–º —Ñ–æ—Ä–º—É–ª—É –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ ROI: (Profit / Cena min.)
                        # –§–æ—Ä–º–∞—Ç 0% –≤ config.ROI_NUMBER_FORMAT –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —É–º–Ω–æ–∂–∏—Ç –Ω–∞ 100
                        formula = f"=IF({cena_min_col_letter}{row_num}<>0,{profit_col_letter}{row_num}/{cena_min_col_letter}{row_num},0)"
                        cell.value = formula
                        
                        # –ü—Ä–∏–º–µ–Ω—è–µ–º —Ñ–æ—Ä–º–∞—Ç —á–∏—Å–ª–∞ (–±—É–¥–µ—Ç –ø—Ä–∏–º–µ–Ω–µ–Ω –ø–æ–∑–∂–µ —á–µ—Ä–µ–∑ config.ROI_NUMBER_FORMAT)
                
                # –û—Å–æ–±–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –¥–ª—è –≥–∏–ø–µ—Ä—Å—Å—ã–ª–æ–∫
                elif column_name == 'Link' and pd.notna(df.iloc[row_num - config.EXCEL_DATA_START_ROW_NUM, col_num - 1]):
                    link_value = df.iloc[row_num - config.EXCEL_DATA_START_ROW_NUM, col_num - 1]
                    if str(link_value).startswith('http'):
                        # –°–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫—É —Å —Ç–µ–∫—Å—Ç–æ–º "Link"
                        cell.hyperlink = str(link_value)
                        cell.value = "Link"  # –û—Ç–æ–±—Ä–∞–∂–∞–µ–º —Ç–µ–∫—Å—Ç "Link" –≤–º–µ—Å—Ç–æ –ø–æ–ª–Ω–æ–π —Å—Å—ã–ª–∫–∏
                        cell.font = Font(name='Arial', size=10, color='0000FF', underline='single')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # –û—Å–æ–±–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –¥–ª—è Product Link –≥–∏–ø–µ—Ä—Å—Å—ã–ª–æ–∫
                elif column_name == 'Product Link' and pd.notna(df.iloc[row_num - config.EXCEL_DATA_START_ROW_NUM, col_num - 1]):
                    product_link_value = df.iloc[row_num - config.EXCEL_DATA_START_ROW_NUM, col_num - 1]
                    if str(product_link_value).startswith('http'):
                        # –°–æ–∑–¥–∞–µ–º –æ–±—ã—á–Ω—É—é –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫—É —á–µ—Ä–µ–∑ openpyxl (–±–æ–ª–µ–µ –Ω–∞–¥–µ–∂–Ω–æ —á–µ–º —Ñ–æ—Ä–º—É–ª–∞)
                        ean_value = df.iloc[row_num - config.EXCEL_DATA_START_ROW_NUM, df.columns.get_loc('EAN')] if 'EAN' in df.columns else ''
                        formatted_ean = format_ean_to_13_digits(ean_value) if ean_value else ''
                        
                        if formatted_ean:
                            # –°–æ–∑–¥–∞–µ–º URL –∏ –ø—Ä–∏—Å–≤–∞–∏–≤–∞–µ–º –∫–∞–∫ –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫—É
                            product_url = f"https://api.qogita.com/variants/link/{formatted_ean}/"
                            cell.hyperlink = product_url
                            cell.value = "View Product"  # –û—Ç–æ–±—Ä–∞–∂–∞–µ–º—ã–π —Ç–µ–∫—Å—Ç
                            cell.font = Font(name='Arial', size=10, color='0000FF', underline='single')
                            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # –î–æ–±–∞–≤–ª—è–µ–º "Date:" –≤ —è—á–µ–π–∫—É A1
        date_label_cell = ws['A1']
        date_label_cell.value = "Date:"
        date_label_cell.font = Font(
            name=config.TITLE_FONT_SETTINGS['name'],
            size=config.TITLE_FONT_SETTINGS['size'],
            bold=config.TITLE_FONT_SETTINGS['bold']
        )
        
        # –î–æ–±–∞–≤–ª—è–µ–º –¥–∞—Ç—É –≤ —è—á–µ–π–∫—É B1
        date_cell = ws['B1']
        date_cell.value = datetime.now().strftime('%d.%m.%Y')
        date_cell.font = Font(
            name=config.TITLE_FONT_SETTINGS['name'],
            size=config.TITLE_FONT_SETTINGS['size'],
            bold=config.TITLE_FONT_SETTINGS['bold']
        )
        
        # –î–æ–±–∞–≤–ª—è–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫—É Help –≤ —è—á–µ–π–∫—É A3
        help_cell = ws['A3']
        help_cell.value = "Help"
        help_cell.hyperlink = "https://t.me/iilluummiinnaattoorr"
        help_cell.font = Font(
            name=config.TITLE_FONT_SETTINGS['name'],
            size=config.TITLE_FONT_SETTINGS['size'],
            bold=config.TITLE_FONT_SETTINGS['bold'],
            color='0000FF',  # –°–∏–Ω–∏–π —Ü–≤–µ—Ç –¥–ª—è –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
            underline='single'
        )
        
        # –ê–∫—Ç–∏–≤–∏—Ä—É–µ–º —Ñ–∏–ª—å—Ç—Ä –Ω–∞ —Å—Ç—Ä–æ–∫–µ –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤ (4-—è —Å—Ç—Ä–æ–∫–∞)
        header_row = config.EXCEL_HEADER_ROW_NUM
        last_column = len(df.columns)
        last_data_row = len(df) + config.EXCEL_DATA_START_ROW_NUM - 1
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –¥–∏–∞–ø–∞–∑–æ–Ω –¥–ª—è —Ñ–∏–ª—å—Ç—Ä–∞ (–æ—Ç –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤ –¥–æ –ø–æ—Å–ª–µ–¥–Ω–µ–π —Å—Ç—Ä–æ–∫–∏ –¥–∞–Ω–Ω—ã—Ö)
        filter_range = f"A{header_row}:{chr(65 + last_column - 1)}{last_data_row}"
        ws.auto_filter.ref = filter_range
        
        # –ó–∞–∫—Ä–µ–ø–ª—è–µ–º –æ–±–ª–∞—Å—Ç–∏ –¥–æ 4-–π —Å—Ç—Ä–æ–∫–∏ –≤–∫–ª—é—á–∏—Ç–µ–ª—å–Ω–æ
        ws.freeze_panes = f"A{config.EXCEL_DATA_START_ROW_NUM}"  # –ó–∞–∫—Ä–µ–ø–ª—è–µ–º –¥–æ 5-–π —Å—Ç—Ä–æ–∫–∏ (–ø–æ—Å–ª–µ –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤)
        
        # –ü—Ä–∏–º–µ–Ω—è–µ–º —É—Å–ª–æ–≤–Ω–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ
        for column_name, format_config in config.CONDITIONAL_FORMAT_COLUMNS.items():
            if column_name in df.columns:
                # –ù–∞—Ö–æ–¥–∏–º –Ω–æ–º–µ—Ä –∫–æ–ª–æ–Ω–∫–∏
                col_index = df.columns.get_loc(column_name) + 1
                col_letter = get_column_letter(col_index)
                
                # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –¥–∏–∞–ø–∞–∑–æ–Ω –¥–∞–Ω–Ω—ã—Ö –¥–ª—è —É—Å–ª–æ–≤–Ω–æ–≥–æ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
                start_row = config.EXCEL_DATA_START_ROW_NUM
                end_row = len(df) + config.EXCEL_DATA_START_ROW_NUM - 1
                range_string = f"{col_letter}{start_row}:{col_letter}{end_row}"
                
                # –°–æ–∑–¥–∞–µ–º –ø—Ä–∞–≤–∏–ª–æ —Ü–≤–µ—Ç–æ–≤–æ–π —à–∫–∞–ª—ã
                rule = ColorScaleRule(
                    start_type='num',
                    start_value=format_config['start_value'],
                    start_color=format_config['start_color'],
                    mid_type='num',
                    mid_value=format_config['mid_value'],
                    mid_color=format_config['mid_color'],
                    end_type='num',
                    end_value=format_config['end_value'],
                    end_color=format_config['end_color']
                )
                
                # –ü—Ä–∏–º–µ–Ω—è–µ–º –ø—Ä–∞–≤–∏–ª–æ –∫ –¥–∏–∞–ø–∞–∑–æ–Ω—É
                ws.conditional_formatting.add(range_string, rule)
                print(f"–ü—Ä–∏–º–µ–Ω–µ–Ω–æ —É—Å–ª–æ–≤–Ω–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –¥–ª—è –∫–æ–ª–æ–Ω–∫–∏ '{column_name}' –≤ –¥–∏–∞–ø–∞–∑–æ–Ω–µ {range_string}")
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ñ–∞–π–ª
        wb.save(output_file)
        print(f"–§–∞–π–ª —Å–æ—Ö—Ä–∞–Ω–µ–Ω —Å —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ–º: {output_file}")
        
    except Exception as e:
        print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–∏ —Ñ–∞–π–ª–∞ {output_file}: {str(e)}")
        # –ï—Å–ª–∏ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –Ω–µ —É–¥–∞–ª–æ—Å—å, —Å–æ—Ö—Ä–∞–Ω—è–µ–º –±–µ–∑ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è
        df.to_excel(output_file, index=False)
        print(f"–§–∞–π–ª —Å–æ—Ö—Ä–∞–Ω–µ–Ω –±–µ–∑ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è: {output_file}")

def add_price_pl_column(df):
    """
    –î–æ–±–∞–≤–ª—è–µ—Ç –∫–æ–ª–æ–Ω–∫—É 'Price PL' –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ —Ü–µ–Ω—ã –ø–æ –∫—É—Ä—Å—É –≤–∞–ª—é—Ç
    """
    # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É Price PL —Å –ø—É—Å—Ç—ã–º–∏ –∑–Ω–∞—á–µ–Ω–∏—è–º–∏
    # –§–æ—Ä–º—É–ª–∞ –±—É–¥–µ—Ç –¥–æ–±–∞–≤–ª–µ–Ω–∞ –ø–æ–∑–∂–µ –ø—Ä–∏ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–∏ Excel
    df['Price PL'] = None
    
    print("–î–æ–±–∞–≤–ª–µ–Ω–∞ –∫–æ–ª–æ–Ω–∫–∞ 'Price PL' –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ —Ü–µ–Ω—ã –ø–æ –∫—É—Ä—Å—É –≤–∞–ª—é—Ç")
    return df

def add_roi_column(df):
    """
    –î–æ–±–∞–≤–ª—è–µ—Ç –∫–æ–ª–æ–Ω–∫—É 'ROI' –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ –≤–æ–∑–≤—Ä–∞—Ç–∞ –∏–Ω–≤–µ—Å—Ç–∏—Ü–∏–π
    """
    # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É ROI —Å –ø—É—Å—Ç—ã–º–∏ –∑–Ω–∞—á–µ–Ω–∏—è–º–∏
    # –§–æ—Ä–º—É–ª–∞ –±—É–¥–µ—Ç –¥–æ–±–∞–≤–ª–µ–Ω–∞ –ø–æ–∑–∂–µ –ø—Ä–∏ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–∏ Excel
    df['ROI'] = None
    
    print("–î–æ–±–∞–≤–ª–µ–Ω–∞ –∫–æ–ª–æ–Ω–∫–∞ 'ROI' –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ –≤–æ–∑–≤—Ä–∞—Ç–∞ –∏–Ω–≤–µ—Å—Ç–∏—Ü–∏–π")
    return df

def merge_excel_files_from_list(file_paths, original_filename=None):
    """
    –û–±—ä–µ–¥–∏–Ω—è–µ—Ç —Ñ–∞–π–ª—ã Excel –ø–æ EAN –∫–æ–¥—É –∏–∑ —Å–ø–∏—Å–∫–∞ —Ñ–∞–π–ª–æ–≤.
    –ü—Ä–µ–¥–Ω–∞–∑–Ω–∞—á–µ–Ω–æ –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã–º–∏ –≤ –±–æ—Ç —Ñ–∞–π–ª–∞–º–∏
    
    Args:
        file_paths: —Å–ø–∏—Å–æ–∫ –ø—É—Ç–µ–π –∫ —Ñ–∞–π–ª–∞–º –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏
        original_filename: –æ—Ä–∏–≥–∏–Ω–∞–ª—å–Ω–æ–µ –∏–º—è —Ñ–∞–π–ª–∞ –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞ (–¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –∏—Ç–æ–≥–æ–≤–æ–≥–æ —Ñ–∞–π–ª–∞)
    
    Returns:
        dict: —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏
    """
    
    # –†–∞–∑–¥–µ–ª—è–µ–º —Ñ–∞–π–ª—ã –Ω–∞ TradeWatch –∏ –æ—Å—Ç–∞–ª—å–Ω—ã–µ
    tradewatch_files = []
    other_files = []
    
    for file_path in file_paths:
        filename = os.path.basename(file_path)
        if filename.startswith('TradeWatch') and filename.endswith('.xlsx'):
            tradewatch_files.append(file_path)
        elif filename.endswith('.xlsx') and not any(filename.startswith(prefix) for prefix in config.EXCLUDED_FILE_PREFIXES):
            other_files.append(file_path)
    
    if not tradewatch_files:
        print("–ù–µ –Ω–∞–π–¥–µ–Ω—ã —Ñ–∞–π–ª—ã TradeWatch —Å—Ä–µ–¥–∏ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã—Ö —Ñ–∞–π–ª–æ–≤")
        return None
    
    print(f"–ù–∞–π–¥–µ–Ω–æ —Ñ–∞–π–ª–æ–≤ TradeWatch: {len(tradewatch_files)}")
    for file in tradewatch_files:
        print(f"  - {os.path.basename(file)}")
    
    # –°–æ–±–∏—Ä–∞–µ–º –≤—Å–µ EAN –∫–æ–¥—ã –∏–∑ —Ñ–∞–π–ª–æ–≤ TradeWatch
    all_ean_data = []
    
    for file_path in tradewatch_files:
        try:
            print(f"\n–û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª: {os.path.basename(file_path)}")
            df = pd.read_excel(file_path, sheet_name=config.TRADEWATCH_SHEET_NAME)
            df['source_file'] = os.path.basename(file_path)
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º EAN –≤ 13-—Ü–∏—Ñ—Ä–æ–≤–æ–º —Ñ–æ—Ä–º–∞—Ç–µ
            df['EAN'] = df['EAN'].apply(format_ean_to_13_digits)
            
            all_ean_data.append(df)
            print(f"  –ù–∞–π–¥–µ–Ω–æ EAN –∫–æ–¥–æ–≤: {len(df)}")
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —á—Ç–µ–Ω–∏–∏ —Ñ–∞–π–ª–∞ {file_path}: {str(e)}")
    
    if not all_ean_data:
        print("–ù–µ —É–¥–∞–ª–æ—Å—å –∑–∞–≥—Ä—É–∑–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –∏–∑ —Ñ–∞–π–ª–æ–≤ TradeWatch")
        return None
    
    # –û–±—ä–µ–¥–∏–Ω—è–µ–º –≤—Å–µ –¥–∞–Ω–Ω—ã–µ TradeWatch
    combined_tradewatch = pd.concat(all_ean_data, ignore_index=True)
    print(f"\n–í—Å–µ–≥–æ —É–Ω–∏–∫–∞–ª—å–Ω—ã—Ö EAN –∫–æ–¥–æ–≤ –∏–∑ TradeWatch: {combined_tradewatch['EAN'].nunique()}")
    
    if not other_files:
        print("–ù–µ –Ω–∞–π–¥–µ–Ω—ã –¥—Ä—É–≥–∏–µ Excel —Ñ–∞–π–ª—ã –¥–ª—è –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏—è")
        # –†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º –ø—Ä–∏–±—ã–ª—å –∏ ROI –¥–ª—è –¥–∞–Ω–Ω—ã—Ö TradeWatch
        combined_tradewatch = calculate_profit_and_roi(combined_tradewatch)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É Price PL
        combined_tradewatch = add_price_pl_column(combined_tradewatch)
        
        # –°–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        combined_tradewatch = create_hyperlinks(combined_tradewatch)
        
        # –°–æ–∑–¥–∞–µ–º Product Link –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        combined_tradewatch = create_product_links(combined_tradewatch)
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –≤—ã—Ö–æ–¥–Ω–æ–π —Ñ–∞–π–ª –≤ —Ç–æ–π –∂–µ –ø–∞–ø–∫–µ, —á—Ç–æ –∏ –ø–µ—Ä–≤—ã–π TradeWatch —Ñ–∞–π–ª
        output_dir = os.path.dirname(tradewatch_files[0])
        output_file = os.path.join(output_dir, config.OUTPUT_FILE_MERGED_TRADEWATCH)
        save_formatted_excel(combined_tradewatch, output_file)
        
        stats = {
            'total_rows': len(combined_tradewatch),
            'unique_ean': combined_tradewatch['EAN'].nunique() if 'EAN' in combined_tradewatch.columns else 0,
            'files_processed': 0,  # —Ç–æ–ª—å–∫–æ TradeWatch —Ñ–∞–π–ª—ã
            'output_file': output_file
        }
        
        print(f"–°–æ—Ö—Ä–∞–Ω–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ TradeWatch –≤ —Ñ–∞–π–ª: {output_file}")
        return stats
    
    print(f"\n–ù–∞–π–¥–µ–Ω–æ –¥—Ä—É–≥–∏—Ö Excel —Ñ–∞–π–ª–æ–≤: {len(other_files)}")
    for file in other_files:
        print(f"  - {os.path.basename(file)}")
    
    # –û–±—ä–µ–¥–∏–Ω—è–µ–º —Å –¥—Ä—É–≥–∏–º–∏ —Ñ–∞–π–ª–∞–º–∏
    merged_results = []
    
    for other_file in other_files:
        try:
            print(f"\n–û–±—ä–µ–¥–∏–Ω—è–µ–º —Å —Ñ–∞–π–ª–æ–º: {os.path.basename(other_file)}")
            
            # –ß–∏—Ç–∞–µ–º —Ñ–∞–π–ª
            other_df = pd.read_excel(other_file)
            
            # –ò—â–µ–º –∫–æ–ª–æ–Ω–∫—É —Å GTIN
            gtin_column = None
            for col in config.POSSIBLE_GTIN_COLUMNS:
                if col in other_df.columns:
                    gtin_column = col
                    break
            
            if gtin_column is None:
                print(f"  –í —Ñ–∞–π–ª–µ {other_file} –Ω–µ –Ω–∞–π–¥–µ–Ω–∞ –∫–æ–ª–æ–Ω–∫–∞ GTIN")
                continue
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º GTIN –≤ 13-—Ü–∏—Ñ—Ä–æ–≤–æ–π —Ñ–æ—Ä–º–∞—Ç
            other_df_clean = other_df[other_df[gtin_column].notna()].copy()
            other_df_clean[gtin_column] = other_df_clean[gtin_column].apply(format_ean_to_13_digits)
            other_df_clean = other_df_clean[other_df_clean[gtin_column].notna()].copy()
            
            # –û–±—ä–µ–¥–∏–Ω—è–µ–º –ø–æ EAN
            merged = pd.merge(combined_tradewatch, other_df_clean, left_on='EAN', right_on=gtin_column, how='inner')
            
            if not merged.empty:
                merged['merged_with'] = os.path.basename(other_file)
                merged_results.append(merged)
                print(f"  –ù–∞–π–¥–µ–Ω–æ —Å–æ–≤–ø–∞–¥–µ–Ω–∏–π: {len(merged)}")
            else:
                print(f"  –°–æ–≤–ø–∞–¥–µ–Ω–∏–π –Ω–µ –Ω–∞–π–¥–µ–Ω–æ")
                
        except Exception as e:
            print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏–∏ —Å —Ñ–∞–π–ª–æ–º {other_file}: {str(e)}")
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
    if merged_results:
        # –û–±—ä–µ–¥–∏–Ω—è–µ–º –≤—Å–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—ã
        final_result = pd.concat(merged_results, ignore_index=True)
        
        # –†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º –ø—Ä–∏–±—ã–ª—å –∏ ROI
        final_result = calculate_profit_and_roi(final_result)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É Price PL
        final_result = add_price_pl_column(final_result)
        
        # –°–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        final_result = create_hyperlinks(final_result)
        
        # –°–æ–∑–¥–∞–µ–º Product Link –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        final_result = create_product_links(final_result)
        
        # –£–ø–æ—Ä—è–¥–æ—á–∏–≤–∞–µ–º –∫–æ–ª–æ–Ω–∫–∏ –≤ –Ω—É–∂–Ω–æ–º –ø–æ—Ä—è–¥–∫–µ –∏ –æ—Å—Ç–∞–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ –Ω—É–∂–Ω—ã–µ
        available_ordered_columns = [col for col in config.DESIRED_COLUMN_ORDER if col in final_result.columns]
        final_column_order = available_ordered_columns
        final_result = final_result[final_column_order]
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –≤—ã—Ö–æ–¥–Ω–æ–π —Ñ–∞–π–ª –≤ —Ç–æ–π –∂–µ –ø–∞–ø–∫–µ, —á—Ç–æ –∏ –ø–µ—Ä–≤—ã–π TradeWatch —Ñ–∞–π–ª
        output_dir = os.path.dirname(tradewatch_files[0])
        
        # –°–æ–∑–¥–∞–µ–º –∏–º—è —Ñ–∞–π–ª–∞ –Ω–∞ –æ—Å–Ω–æ–≤–µ –æ—Ä–∏–≥–∏–Ω–∞–ª—å–Ω–æ–≥–æ –∏–º–µ–Ω–∏ —Å timestamp
        if original_filename:
            # –£–±–∏—Ä–∞–µ–º —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ –∏ –¥–æ–±–∞–≤–ª—è–µ–º timestamp
            base_name = os.path.splitext(os.path.basename(original_filename))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"{base_name}_result_{timestamp}.xlsx"
        else:
            # –ò—Å–ø–æ–ª—å–∑—É–µ–º —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω–æ–µ –∏–º—è, –µ—Å–ª–∏ –æ—Ä–∏–≥–∏–Ω–∞–ª—å–Ω–æ–µ –∏–º—è –Ω–µ –ø—Ä–µ–¥–æ—Å—Ç–∞–≤–ª–µ–Ω–æ
            output_filename = config.OUTPUT_FILE_WITH_CALCULATIONS
        
        output_file = os.path.join(output_dir, output_filename)
        save_formatted_excel(final_result, output_file)
        
        # –ü–æ–¥–≥–æ—Ç–∞–≤–ª–∏–≤–∞–µ–º —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫—É
        stats = {
            'total_rows': len(final_result),
            'unique_ean': final_result['EAN'].nunique() if 'EAN' in final_result.columns else 0,
            'files_processed': len(other_files),
            'output_file': output_file
        }
        
        print(f"\n–†–µ–∑—É–ª—å—Ç–∞—Ç —Å–æ—Ö—Ä–∞–Ω–µ–Ω –≤ —Ñ–∞–π–ª: {output_file}")
        print(f"–í—Å–µ–≥–æ —Å—Ç—Ä–æ–∫ –≤ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–µ: {stats['total_rows']}")
        print(f"–£–Ω–∏–∫–∞–ª—å–Ω—ã—Ö EAN –∫–æ–¥–æ–≤: {stats['unique_ean']}")
        print(f"–ü–æ—Ä—è–¥–æ–∫ –∫–æ–ª–æ–Ω–æ–∫: {final_column_order}")
        
        return stats
    else:
        print("\n–°–æ–≤–ø–∞–¥–µ–Ω–∏–π –Ω–µ –Ω–∞–π–¥–µ–Ω–æ, —Å–æ—Ö—Ä–∞–Ω—è–µ–º —Ç–æ–ª—å–∫–æ –¥–∞–Ω–Ω—ã–µ TradeWatch")
        # –†–∞—Å—Å—á–∏—Ç—ã–≤–∞–µ–º –ø—Ä–∏–±—ã–ª—å –∏ ROI –¥–ª—è –¥–∞–Ω–Ω—ã—Ö TradeWatch
        combined_tradewatch = calculate_profit_and_roi(combined_tradewatch)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É Price PL
        combined_tradewatch = add_price_pl_column(combined_tradewatch)
        
        # –°–æ–∑–¥–∞–µ–º –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        combined_tradewatch = create_hyperlinks(combined_tradewatch)
        
        # –°–æ–∑–¥–∞–µ–º Product Link –≥–∏–ø–µ—Ä—Å—Å—ã–ª–∫–∏
        combined_tradewatch = create_product_links(combined_tradewatch)
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –≤—ã—Ö–æ–¥–Ω–æ–π —Ñ–∞–π–ª –≤ —Ç–æ–π –∂–µ –ø–∞–ø–∫–µ, —á—Ç–æ –∏ –ø–µ—Ä–≤—ã–π TradeWatch —Ñ–∞–π–ª
        output_dir = os.path.dirname(tradewatch_files[0])
        output_file = os.path.join(output_dir, config.OUTPUT_FILE_TRADEWATCH_ONLY)
        save_formatted_excel(combined_tradewatch, output_file)
        
        stats = {
            'total_rows': len(combined_tradewatch),
            'unique_ean': combined_tradewatch['EAN'].nunique() if 'EAN' in combined_tradewatch.columns else 0,
            'files_processed': 0,
            'output_file': output_file
        }
        
        print(f"–°–æ—Ö—Ä–∞–Ω–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ TradeWatch –≤ —Ñ–∞–π–ª: {output_file}")
        return stats

def process_supplier_with_tradewatch_auto(supplier_file_path, temp_dir, progress_callback=None):
    """
    –ù–æ–≤–∞—è —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ñ–∞–π–ª–∞ –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞ —Å TradeWatch
    
    Args:
        supplier_file_path: –ø—É—Ç—å –∫ —Ñ–∞–π–ª—É –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞
        temp_dir: –≤—Ä–µ–º–µ–Ω–Ω–∞—è –ø–∞–ø–∫–∞ –¥–ª—è —Å–∫–∞—á–∏–≤–∞–Ω–∏—è —Ñ–∞–π–ª–æ–≤
        progress_callback: —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –æ—Ç—Å–ª–µ–∂–∏–≤–∞–Ω–∏—è –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ (–æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–æ)
    
    Returns:
        dict: —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∏ –ø—É—Ç—å –∫ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—É
    """
    try:
        print(f"–ù–∞—á–∏–Ω–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É —Ñ–∞–π–ª–∞ –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞: {supplier_file_path}")
        
        # –°–æ–∑–¥–∞–µ–º –≤—Ä–µ–º–µ–Ω–Ω—É—é –ø–∞–ø–∫—É –¥–ª—è —Å–∫–∞—á–∏–≤–∞–Ω–∏—è
        download_dir = os.path.join(temp_dir, "tradewatch_downloads")
        os.makedirs(download_dir, exist_ok=True)
        
        # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞ –∏ –ø–æ–ª—É—á–∞–µ–º —Ñ–∞–π–ª—ã TradeWatch
        print("–ò–∑–≤–ª–µ–∫–∞–µ–º EAN –∫–æ–¥—ã –∏ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —á–µ—Ä–µ–∑ TradeWatch...")
        
        if SELENIUM_AVAILABLE:
            tradewatch_files = process_supplier_file_with_tradewatch(supplier_file_path, download_dir, progress_callback=progress_callback)
        else:
            # Fallback —Ä–µ–∂–∏–º - –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –ø—É—Å—Ç–æ–π —Ä–µ–∑—É–ª—å—Ç–∞—Ç —Å –∏–Ω—Ñ–æ—Ä–º–∞—Ç–∏–≤–Ω—ã–º —Å–æ–æ–±—â–µ–Ω–∏–µ–º
            if progress_callback:
                progress_callback("‚ùå TradeWatch –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω –±–µ–∑ Selenium")
            
            return {
                'success': False,
                'error': 'TradeWatch –∏–Ω—Ç–µ–≥—Ä–∞—Ü–∏—è –Ω–µ–¥–æ—Å—Ç—É–ø–Ω–∞ –±–µ–∑ Selenium.\n–ë–æ—Ç —Ä–∞–±–æ—Ç–∞–µ—Ç –≤ –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–Ω–æ–º —Ä–µ–∂–∏–º–µ - –º–æ–∂–µ—Ç–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å —Ç–æ–ª—å–∫–æ –æ–±—Ä–∞–±–æ—Ç–∫—É Excel —Ñ–∞–π–ª–æ–≤ –±–µ–∑ –∞–Ω–∞–ª–∏–∑–∞ –∫–æ–Ω–∫—É—Ä–µ–Ω—Ç–æ–≤.',
                'message': '–î–ª—è –ø–æ–ª–Ω–æ–π —Ñ—É–Ω–∫—Ü–∏–æ–Ω–∞–ª—å–Ω–æ—Å—Ç–∏ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ —Ä–∞–∑–≤–µ—Ä—Ç—ã–≤–∞–Ω–∏–µ —Å Selenium'
            }
        
        if not tradewatch_files:
            return {
                'success': False,
                'error': '–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –∏–∑ TradeWatch',
                'files_processed': 0
            }
        
        print(f"–ü–æ–ª—É—á–µ–Ω–æ {len(tradewatch_files)} —Ñ–∞–π–ª–æ–≤ TradeWatch")
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –≤—Å–µ —Ñ–∞–π–ª—ã —Å—É—â–µ—Å—Ç–≤—É—é—Ç
        print("–ü—Ä–æ–≤–µ—Ä–∫–∞ —Ñ–∞–π–ª–æ–≤ TradeWatch:")
        for file_path in tradewatch_files:
            if os.path.exists(file_path):
                size = os.path.getsize(file_path)
                print(f"  ‚úÖ {file_path} (—Ä–∞–∑–º–µ—Ä: {size} –±–∞–π—Ç)")
            else:
                print(f"  ‚ùå {file_path} - –ù–ï –ù–ê–ô–î–ï–ù!")
        
        # –û–±—ä–µ–¥–∏–Ω—è–µ–º —Ñ–∞–π–ª—ã –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞ —Å —Ñ–∞–π–ª–∞–º–∏ TradeWatch
        print("–û–±—ä–µ–¥–∏–Ω—è–µ–º —Ñ–∞–π–ª—ã...")
        
        # –°–æ–∑–¥–∞–µ–º —Å–ø–∏—Å–æ–∫ –≤—Å–µ—Ö —Ñ–∞–π–ª–æ–≤ –¥–ª—è –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏—è
        all_files = [supplier_file_path] + tradewatch_files
        result = merge_excel_files_from_list(all_files, supplier_file_path)
        
        if result:
            print(f"–û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞ —É—Å–ø–µ—à–Ω–æ!")
            print(f"–†–µ–∑—É–ª—å—Ç–∞—Ç —Å–æ—Ö—Ä–∞–Ω–µ–Ω –≤: {result['output_file']}")
            
            # –ù–ï —É–¥–∞–ª—è–µ–º –≤—Ä–µ–º–µ–Ω–Ω—ã–µ —Ñ–∞–π–ª—ã TradeWatch —Å—Ä–∞–∑—É - –æ–Ω–∏ –ø–æ–Ω–∞–¥–æ–±—è—Ç—Å—è –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
            # –û—á–∏—Å—Ç–∫–∞ –ø—Ä–æ–∏—Å—Ö–æ–¥–∏—Ç –≤ telegram_bot.py –ø–æ—Å–ª–µ –æ—Ç–ø—Ä–∞–≤–∫–∏ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞
            
            return {
                'success': True,
                'output_file': result['output_file'],
                'total_rows': result['total_rows'],
                'unique_ean': result['unique_ean'],
                'files_processed': len(tradewatch_files),
                'supplier_file': supplier_file_path,
                'tradewatch_files_count': len(tradewatch_files)
            }
        else:
            return {
                'success': False,
                'error': '–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏–∏ —Ñ–∞–π–ª–æ–≤',
                'files_processed': len(tradewatch_files)
            }
            
    except Exception as e:
        print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'files_processed': 0
        }

def main():
    """–û—Å–Ω–æ–≤–Ω–∞—è —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –∑–∞–ø—É—Å–∫–∞ —Å–∫—Ä–∏–ø—Ç–∞ —Å —Ä–∞—Å—á–µ—Ç–∞–º–∏"""
    current_directory = os.getcwd()
    print(f"–†–∞–±–æ—á–∞—è –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏—è: {current_directory}")
    print("=" * 50)
    
    merge_excel_files_by_ean_with_calculations(current_directory)

if __name__ == "__main__":
    main()


def process_supplier_with_tradewatch_interruptible(supplier_file_path, temp_dir, stop_flag_callback=None, progress_callback=None):
    """
    –§—É–Ω–∫—Ü–∏—è –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ñ–∞–π–ª–∞ –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞ —Å –≤–æ–∑–º–æ–∂–Ω–æ—Å—Ç—å—é –æ—Å—Ç–∞–Ω–æ–≤–∫–∏ –ø—Ä–æ—Ü–µ—Å—Å–∞
    
    Args:
        supplier_file_path: –ø—É—Ç—å –∫ —Ñ–∞–π–ª—É –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞
        temp_dir: –≤—Ä–µ–º–µ–Ω–Ω–∞—è –ø–∞–ø–∫–∞ –¥–ª—è —Å–∫–∞—á–∏–≤–∞–Ω–∏—è —Ñ–∞–π–ª–æ–≤
        stop_flag_callback: —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ —Ñ–ª–∞–≥–∞ –æ—Å—Ç–∞–Ω–æ–≤–∫–∏
        progress_callback: —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –ø—Ä–æ–≥—Ä–µ—Å—Å–∞
    
    Returns:
        dict: —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∏ –ø—É—Ç—å –∫ —Ä–µ–∑—É–ª—å—Ç–∞—Ç—É
    """
    try:
        print(f"–ù–∞—á–∏–Ω–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É —Ñ–∞–π–ª–∞ –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞: {supplier_file_path}")
        
        # –°–æ–∑–¥–∞–µ–º –≤—Ä–µ–º–µ–Ω–Ω—É—é –ø–∞–ø–∫—É –¥–ª—è —Å–∫–∞—á–∏–≤–∞–Ω–∏—è
        download_dir = os.path.join(temp_dir, "tradewatch_downloads")
        os.makedirs(download_dir, exist_ok=True)
        
        # –ò–º–ø–æ—Ä—Ç–∏—Ä—É–µ–º —Ñ—É–Ω–∫—Ü–∏—é —Å –ø–æ–¥–¥–µ—Ä–∂–∫–æ–π –æ—Å—Ç–∞–Ω–æ–≤–∫–∏
        from tradewatch_login import process_supplier_file_with_tradewatch_interruptible
        
        # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞ –∏ –ø–æ–ª—É—á–∞–µ–º —Ñ–∞–π–ª—ã TradeWatch
        if progress_callback:
            progress_callback("üîÑ –ò–∑–≤–ª–µ–∫–∞—é EAN –∫–æ–¥—ã –∏ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞—é —á–µ—Ä–µ–∑ TradeWatch...")
        
        print("–ò–∑–≤–ª–µ–∫–∞–µ–º EAN –∫–æ–¥—ã –∏ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —á–µ—Ä–µ–∑ TradeWatch...")
        tradewatch_files = process_supplier_file_with_tradewatch_interruptible(
            supplier_file_path, 
            download_dir, 
            stop_flag_callback=stop_flag_callback,
            progress_callback=progress_callback
        )
        
        if not tradewatch_files:
            return {
                'success': False,
                'error': '–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –∏–∑ TradeWatch',
                'files_processed': 0
            }
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ñ–ª–∞–≥ –æ—Å—Ç–∞–Ω–æ–≤–∫–∏ –ø–µ—Ä–µ–¥ –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏–µ–º
        if stop_flag_callback and stop_flag_callback():
            print("üõë –ü—Ä–æ—Ü–µ—Å—Å –æ—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω –ø–µ—Ä–µ–¥ –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏–µ–º —Ñ–∞–π–ª–æ–≤")
        
        print(f"–ü–æ–ª—É—á–µ–Ω–æ {len(tradewatch_files)} —Ñ–∞–π–ª–æ–≤ TradeWatch")
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –≤—Å–µ —Ñ–∞–π–ª—ã —Å—É—â–µ—Å—Ç–≤—É—é—Ç
        print("–ü—Ä–æ–≤–µ—Ä–∫–∞ —Ñ–∞–π–ª–æ–≤ TradeWatch:")
        for file_path in tradewatch_files:
            if os.path.exists(file_path):
                size = os.path.getsize(file_path)
                print(f"  ‚úÖ {file_path} (—Ä–∞–∑–º–µ—Ä: {size} –±–∞–π—Ç)")
            else:
                print(f"  ‚ùå {file_path} - –ù–ï –ù–ê–ô–î–ï–ù!")
        
        # –û–±—ä–µ–¥–∏–Ω—è–µ–º —Ñ–∞–π–ª—ã –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞ —Å —Ñ–∞–π–ª–∞–º–∏ TradeWatch
        if progress_callback:
            progress_callback("üìä –û–±—ä–µ–¥–∏–Ω—è—é —Ñ–∞–π–ª—ã –∏ —Å–æ–∑–¥–∞—é –æ—Ç—á—ë—Ç...")
        
        print("–û–±—ä–µ–¥–∏–Ω—è–µ–º —Ñ–∞–π–ª—ã...")
        
        # –°–æ–∑–¥–∞–µ–º —Å–ø–∏—Å–æ–∫ –≤—Å–µ—Ö —Ñ–∞–π–ª–æ–≤ –¥–ª—è –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏—è
        all_files = [supplier_file_path] + tradewatch_files
        result = merge_excel_files_from_list(all_files, supplier_file_path)
        
        if result:
            status_msg = "üõë –ß–∞—Å—Ç–∏—á–Ω—ã–π –æ—Ç—á—ë—Ç —Å–æ–∑–¥–∞–Ω!" if (stop_flag_callback and stop_flag_callback()) else "‚úÖ –û–±—Ä–∞–±–æ—Ç–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞ —É—Å–ø–µ—à–Ω–æ!"
            print(status_msg)
            print(f"–†–µ–∑—É–ª—å—Ç–∞—Ç —Å–æ—Ö—Ä–∞–Ω–µ–Ω –≤: {result['output_file']}")
            
            return {
                'success': True,
                'output_file': result['output_file'],
                'total_rows': result['total_rows'],
                'unique_ean': result['unique_ean'],
                'files_processed': len(tradewatch_files),
                'supplier_file': supplier_file_path,
                'tradewatch_files_count': len(tradewatch_files),
                'is_partial': stop_flag_callback and stop_flag_callback() if stop_flag_callback else False
            }
        else:
            return {
                'success': False,
                'error': '–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏–∏ —Ñ–∞–π–ª–æ–≤',
                'files_processed': len(tradewatch_files)
            }
            
    except Exception as e:
        print(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'files_processed': 0
        }
